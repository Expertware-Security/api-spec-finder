#!/usr/bin/env python3
"""
fast_http_scan.py - masscan-style, pure-Python async port sweep -> HTTP identification.

Authorized internal security assessment tool (defensive / pentest use).

Two phases:
  1. Fast async TCP connect sweep across every requested port on every target host
     (masscan-style: high concurrency, short timeout, RST-close to avoid TIME_WAIT).
  2. HTTP/HTTPS identification on ONLY the ports found open, so the final result is
     "which ports of each host actually speak HTTP".

No external scanner binary is used (no masscan/nmap) - the sweep logic is implemented
here with asyncio. Tuned for Windows (ProactorEventLoop / IOCP, SO_LINGER RST-close),
works on POSIX too.

Outputs (base name from -o, default scan_<timestamp>):
  <base>.urls.txt   scheme://host:port  (one per HTTP endpoint - feed to api_auth_recon.py)
  <base>.byhost.txt human view: only the HTTP ports of each host
  <base>.open.txt   every open ip:port from phase 1 (HTTP or not)
  <base>.http.jsonl full per-endpoint detail
  <base>.http.csv   per-endpoint detail (spreadsheet)
  <base>.xlsx       Excel: "HTTP by host" + "Endpoints" sheets (needs openpyxl)
  <base>.journal.jsonl  crash-safe journal for --resume
"""

import argparse
import asyncio
import csv
import glob
import ipaddress
import json
import math
import os
import random
import re
import socket
import ssl
import struct
import sys
import time
from collections import defaultdict
from datetime import datetime

WIN = sys.platform == "win32"

# onoff=1, linger=0  -> close() sends RST, socket skips TIME_WAIT (portable idiom).
LINGER_ON = struct.pack("ii", 1, 0)

# Approx single-process connect-attempts/sec ceiling for this style of async scan.
# One asyncio loop is CPU-bound well before the network is, so we scale across
# PROCESSES to go faster (each worker is its own loop + GIL). Used only for estimates.
PER_PROC_CPS = 1500

# Ports where TLS is the likely first guess (try https before http to save a round-trip).
TLS_HINT_PORTS = {
    443, 832, 981, 1311, 4433, 4443, 5443, 6443, 7443, 8443, 9443, 10443,
    8834, 9444, 2443, 3443, 4444, 12443, 16443, 18443,
}

DEFAULT_UA = "fast-http-scan/1.0"


# --------------------------------------------------------------------------- #
# Target / port parsing
# --------------------------------------------------------------------------- #
def _is_ip(s):
    try:
        ipaddress.ip_address(str(s))
        return True
    except ValueError:
        return False


_RANGE_RE = re.compile(r"^\d{1,3}(?:\.\d{1,3}){3}-(?:\d{1,3}(?:\.\d{1,3}){3}|\d{1,3})$")


def parse_ports(spec):
    """'1-65535' / 'all' / '80,443,8000-8100' -> sorted unique list, clamped 1..65535."""
    s = spec.strip().lower()
    if s in ("all", "*", "-", "1-65535", "0-65535"):
        return list(range(1, 65536))
    out = set()
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            a, b = int(a), int(b)
            if a > b:
                a, b = b, a
            a, b = max(1, a), min(65535, b)
            out.update(range(a, b + 1))
        else:
            p = int(part)
            if 1 <= p <= 65535:
                out.add(p)
    return sorted(out)


def _expand_range(token):
    """'10.0.0.1-10.0.0.50' or '10.0.0.1-50' -> ip strings."""
    start_s, end_s = token.split("-", 1)
    start = ipaddress.ip_address(start_s.strip())
    end_s = end_s.strip()
    if "." not in end_s:
        base = str(start).rsplit(".", 1)[0]
        end = ipaddress.ip_address(f"{base}.{end_s}")
    else:
        end = ipaddress.ip_address(end_s)
    lo, hi = sorted((int(start), int(end)))
    for i in range(lo, hi + 1):
        yield str(ipaddress.ip_address(i))


def expand_targets(lines):
    """
    Parse target lines into raw (label, host, is_hostname) tuples.
    Accepts: IP, CIDR (10.0.0.0/24), IP range (10.0.0.1-50 / 10.0.0.1-10.0.0.9),
    hostname, or a pasted URL (scheme/path/port are stripped).
    """
    raw = []
    for line in lines:
        s = line.split("#", 1)[0].strip()
        if not s:
            continue
        s = re.sub(r"^\w+://", "", s)          # strip scheme if a URL was pasted
        s = s.split("/", 1)[0] if "/" not in s or _looks_cidr(s) else s
        if not _looks_cidr(s):
            s = s.split("/", 1)[0]             # strip any path
        s = s.strip()
        if not s:
            continue
        if _looks_cidr(s):
            try:
                net = ipaddress.ip_network(s, strict=False)
            except ValueError:
                sys.stderr.write(f"[warn] bad CIDR skipped: {s}\n")
                continue
            iterable = net.hosts() if net.num_addresses > 2 else net
            for ip in iterable:
                raw.append((str(ip), str(ip), False))
            continue
        if _RANGE_RE.match(s):
            for ip in _expand_range(s):
                raw.append((ip, ip, False))
            continue
        # strip :port if someone put host:port
        if s.count(":") == 1 and not _is_ip(s):
            host_part = s.split(":", 1)[0]
        else:
            host_part = s
        if _is_ip(host_part):
            raw.append((host_part, host_part, False))
        else:
            raw.append((host_part, host_part, True))
    return raw


def _looks_cidr(s):
    if "/" not in s:
        return False
    head = s.split("/", 1)[0]
    return _is_ip(head)


def resolve_targets(raw, verbose=True):
    """raw -> deduped list of (label, ip). Hostnames resolved concurrently; ip dedup prefers a hostname label."""
    host_entries = [(l, h) for (l, h, is_host) in raw if is_host]
    ip_entries = [(l, h) for (l, h, is_host) in raw if not is_host]
    resolved = []
    if host_entries:
        from concurrent.futures import ThreadPoolExecutor

        def _r(item):
            label, host = item
            try:
                return (label, socket.gethostbyname(host))
            except OSError:
                return None

        with ThreadPoolExecutor(max_workers=min(64, len(host_entries))) as ex:
            for res in ex.map(_r, host_entries):
                if res:
                    resolved.append(res)
        failed = len(host_entries) - len(resolved)
        if failed and verbose:
            sys.stderr.write(f"[warn] {failed} hostname(s) could not be resolved\n")

    best = {}
    for label, ip in ip_entries + resolved:
        if ip not in best or (_is_ip(best[ip]) and not _is_ip(label)):
            best[ip] = label
    return [(best[ip], ip) for ip in best]


# --------------------------------------------------------------------------- #
# Progress
# --------------------------------------------------------------------------- #
def fmt_dur(sec):
    if sec != sec or sec == float("inf"):
        return "?"
    sec = int(sec)
    h, rem = divmod(sec, 3600)
    m, s = divmod(rem, 60)
    if h:
        return f"{h}h{m:02d}m{s:02d}s"
    if m:
        return f"{m}m{s:02d}s"
    return f"{s}s"


class Progress:
    def __init__(self, total, label, quiet=False):
        self.total = max(1, total)
        self.label = label
        self.quiet = quiet
        self.done = 0
        self.open = 0
        self.start = time.time()
        self._last_t = self.start
        self._last_done = 0
        self._stop = False

    def hit(self, opened=False):
        self.done += 1
        if opened:
            self.open += 1

    async def run(self):
        while not self._stop:
            await asyncio.sleep(2.0)
            self._render()

    def _render(self):
        if self.quiet:
            return
        now = time.time()
        dt = max(1e-6, now - self._last_t)
        rate = (self.done - self._last_done) / dt
        self._last_t, self._last_done = now, self.done
        pct = 100.0 * self.done / self.total
        eta = (self.total - self.done) / rate if rate > 0 else float("inf")
        sys.stdout.write(
            f"\r[{self.label}] {self.done:,}/{self.total:,} ({pct:5.1f}%)  "
            f"{rate:,.0f}/s  open={self.open}  ETA {fmt_dur(eta)}      "
        )
        sys.stdout.flush()

    def finish(self):
        self._stop = True
        if self.quiet:
            return
        el = time.time() - self.start
        sys.stdout.write(
            f"\r[{self.label}] {self.done:,}/{self.total:,} done in {fmt_dur(el)}  "
            f"open={self.open}{' ' * 24}\n"
        )
        sys.stdout.flush()


# --------------------------------------------------------------------------- #
# Journal (crash-safe resume)
# --------------------------------------------------------------------------- #
class Journal:
    def __init__(self, path, enabled=True):
        self.path = path
        self.fh = open(path, "a", buffering=1, encoding="utf-8") if enabled else None

    def open_port(self, label, ip, port):
        if self.fh:
            self.fh.write(json.dumps({"t": "open", "l": label, "ip": ip, "p": port}) + "\n")

    def ports_done(self, ports):
        if self.fh:
            self.fh.write(json.dumps({"t": "pd", "ports": ports}) + "\n")

    def close(self):
        if self.fh:
            self.fh.close()
            self.fh = None


def load_journal(path):
    done_ports, opens = set(), []
    if not os.path.exists(path):
        return done_ports, opens
    with open(path, encoding="utf-8") as f:
        for line in f:
            try:
                r = json.loads(line)
            except ValueError:
                continue
            if r.get("t") == "pd":
                done_ports.update(r.get("ports", []))
            elif r.get("t") == "open":
                opens.append((r["l"], r["ip"], r["p"]))
    return done_ports, opens


def load_all_journals(base):
    """Merge the single-process journal and every per-worker shard journal."""
    done, opens = set(), []
    files = [base + ".journal.jsonl"] + sorted(glob.glob(base + ".journal.p*.jsonl"))
    for path in files:
        d, o = load_journal(path)
        done |= d
        opens += o
    return done, opens


# --------------------------------------------------------------------------- #
# Phase 1: async TCP connect sweep
# --------------------------------------------------------------------------- #
async def scan_open_ports(targets, ports, cfg, journal, hit):
    """Core sweep: connect to every (host, port); call hit(opened) per attempt."""
    loop = asyncio.get_running_loop()
    sem = asyncio.Semaphore(cfg["concurrency"])
    timeout = cfg["connect_timeout"]
    linger = cfg["linger"]
    opens = []

    async def check(label, ip, port):
        async with sem:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            s.setblocking(False)
            try:
                await asyncio.wait_for(loop.sock_connect(s, (ip, port)), timeout)
                if linger:
                    try:
                        s.setsockopt(socket.SOL_SOCKET, socket.SO_LINGER, LINGER_ON)
                    except OSError:
                        pass
                opens.append((label, ip, port))
                journal.open_port(label, ip, port)
                hit(True)
            except (asyncio.TimeoutError, OSError):
                hit(False)
            finally:
                s.close()

    hosts = list(targets)
    random.shuffle(hosts)                       # spread load across hosts, masscan-style
    port_list = list(ports)
    random.shuffle(port_list)
    per_batch_ports = max(1, cfg["pairs_per_batch"] // max(1, len(hosts)))

    for i in range(0, len(port_list), per_batch_ports):
        batch = port_list[i : i + per_batch_ports]
        tasks = [
            asyncio.create_task(check(label, ip, port))
            for port in batch
            for (label, ip) in hosts
        ]
        await asyncio.gather(*tasks)
        journal.ports_done(batch)
    return opens


def run_single(targets, ports, cfg, journal):
    """Single-process sweep with an in-process live progress line."""
    prog = Progress(len(targets) * len(ports), "sweep", cfg["quiet"])

    async def _drive():
        printer = asyncio.create_task(prog.run())
        try:
            return await scan_open_ports(targets, ports, cfg, journal, prog.hit)
        finally:
            prog.finish()
            printer.cancel()
            try:
                await printer
            except asyncio.CancelledError:
                pass

    return asyncio.run(_drive())


class ShardCounter:
    """Per-worker progress that flushes into a shared, lock-free array slot."""
    def __init__(self, arr, idx, flush_every=256):
        self.arr = arr
        self.i0 = 2 * idx
        self.done = 0
        self.open = 0
        self._n = 0
        self._flush_every = flush_every

    def hit(self, opened=False):
        self.done += 1
        if opened:
            self.open += 1
        self._n += 1
        if self._n >= self._flush_every:
            self.arr[self.i0] = self.done
            self.arr[self.i0 + 1] = self.open
            self._n = 0

    def flush(self):
        self.arr[self.i0] = self.done
        self.arr[self.i0 + 1] = self.open


def _worker_entry(idx, targets, ports_shard, cfg, base, counters, result_q):
    """Subprocess entry point: sweep one port shard, report opens back via queue."""
    if WIN:
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    counter = ShardCounter(counters, idx)
    journal = Journal(f"{base}.journal.p{idx}.jsonl", enabled=cfg.get("journal", True))
    opens = []
    try:
        opens = asyncio.run(
            scan_open_ports(targets, ports_shard, cfg, journal, counter.hit)
        )
    except Exception as e:  # keep the parent alive even if one worker dies
        sys.stderr.write(f"[worker {idx}] error: {e}\n")
    finally:
        counter.flush()
        journal.close()
    result_q.put((idx, opens))


def run_multi(targets, ports, cfg, base, nproc):
    """Multi-process sweep: shard ports across nproc workers; parent aggregates."""
    import multiprocessing as mp
    import threading

    ctx = mp.get_context("spawn")
    counters = ctx.Array("Q", 2 * nproc, lock=False)
    result_q = ctx.Queue()

    port_list = list(ports)
    random.shuffle(port_list)
    shards = [[] for _ in range(nproc)]
    for i, p in enumerate(port_list):
        shards[i % nproc].append(p)

    procs = []
    for i in range(nproc):
        pr = ctx.Process(
            target=_worker_entry,
            args=(i, targets, shards[i], cfg, base, counters, result_q),
        )
        pr.start()
        procs.append(pr)

    total = len(targets) * len(ports)
    stop = threading.Event()

    def _printer():
        last, last_t, start = 0, time.time(), time.time()
        while not stop.is_set():
            stop.wait(2.0)
            done = sum(counters[2 * i] for i in range(nproc))
            opn = sum(counters[2 * i + 1] for i in range(nproc))
            now = time.time()
            dt = max(1e-6, now - last_t)
            rate = (done - last) / dt
            last, last_t = done, now
            pct = 100.0 * done / max(1, total)
            eta = (total - done) / rate if rate > 0 else float("inf")
            sys.stdout.write(
                f"\r[sweep] {done:,}/{total:,} ({pct:5.1f}%)  {rate:,.0f}/s  "
                f"open={opn}  ETA {fmt_dur(eta)}  [{nproc}p]      "
            )
            sys.stdout.flush()

    printer = None
    if not cfg["quiet"]:
        printer = threading.Thread(target=_printer, daemon=True)
        printer.start()

    opens = []
    try:
        for _ in range(nproc):
            _, o = result_q.get()          # drain before join to avoid deadlock
            opens.extend(o)
    except KeyboardInterrupt:
        for p in procs:
            p.terminate()
        stop.set()
        raise
    finally:
        stop.set()
        if printer:
            printer.join(timeout=3)

    for p in procs:
        p.join()

    if not cfg["quiet"]:
        done = sum(counters[2 * i] for i in range(nproc))
        opn = sum(counters[2 * i + 1] for i in range(nproc))
        sys.stdout.write(
            f"\r[sweep] {done:,}/{total:,} done  open={opn}  [{nproc}p]{' ' * 24}\n"
        )
        sys.stdout.flush()
    return opens


# --------------------------------------------------------------------------- #
# Phase 2: HTTP / HTTPS identification
# --------------------------------------------------------------------------- #
def _parse_http(label, ip, port, tls, data):
    scheme = "https" if tls else "http"
    text = data.decode("latin-1", "replace")
    head, _, body = text.partition("\r\n\r\n")
    lines = head.split("\r\n")
    status = None
    m = re.match(r"HTTP/\d(?:\.\d)?\s+(\d{3})", lines[0])
    if m:
        status = int(m.group(1))
    hdrs = {}
    for ln in lines[1:]:
        if ":" in ln:
            k, v = ln.split(":", 1)
            hdrs[k.strip().lower()] = v.strip()
    title = ""
    tm = re.search(r"<title[^>]*>(.*?)</title>", body, re.I | re.S)
    if tm:
        title = re.sub(r"\s+", " ", tm.group(1)).strip()[:160]
    return {
        "label": label,
        "ip": ip,
        "port": port,
        "scheme": scheme,
        "url": f"{scheme}://{label}:{port}/",
        "status": status,
        "server": hdrs.get("server", ""),
        "content_type": hdrs.get("content-type", ""),
        "location": hdrs.get("location", ""),
        "title": title,
    }


async def _http_try(label, ip, port, tls, cfg):
    ctx = None
    server_hostname = None
    if tls:
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        try:
            ctx.set_alpn_protocols(["http/1.1"])
        except (NotImplementedError, ssl.SSLError):
            pass
        server_hostname = label if not _is_ip(label) else None
    try:
        fut = asyncio.open_connection(host=ip, port=port, ssl=ctx, server_hostname=server_hostname)
        reader, writer = await asyncio.wait_for(fut, cfg["http_connect_timeout"])
    except (asyncio.TimeoutError, OSError, ssl.SSLError):
        return None
    data = b""
    try:
        req = (
            f"GET / HTTP/1.1\r\nHost: {label}\r\nUser-Agent: {cfg['ua']}\r\n"
            f"Accept: */*\r\nConnection: close\r\n\r\n"
        ).encode("latin-1", "ignore")
        writer.write(req)
        try:
            await asyncio.wait_for(writer.drain(), cfg["http_read_timeout"])
        except (asyncio.TimeoutError, OSError, ssl.SSLError):
            pass
        try:
            data = await asyncio.wait_for(reader.read(65536), cfg["http_read_timeout"])
        except (asyncio.TimeoutError, OSError, ssl.SSLError):
            data = b""
    finally:
        try:
            writer.close()
            await asyncio.wait_for(writer.wait_closed(), 1.0)
        except (asyncio.TimeoutError, OSError, ssl.SSLError):
            pass
    if data[:5] != b"HTTP/":
        return None
    return _parse_http(label, ip, port, tls, data)


async def probe_http(label, ip, port, cfg):
    order = (True, False) if port in TLS_HINT_PORTS else (False, True)
    for tls in order:
        rec = await _http_try(label, ip, port, tls, cfg)
        if rec:
            return rec
    return None


async def identify_http(opens, cfg):
    sem = asyncio.Semaphore(cfg["http_concurrency"])
    results = []
    prog = Progress(len(opens), "http ", cfg["quiet"])
    printer = asyncio.create_task(prog.run())

    async def one(label, ip, port):
        async with sem:
            rec = await probe_http(label, ip, port, cfg)
            prog.hit(opened=bool(rec))
            if rec:
                results.append(rec)

    try:
        await asyncio.gather(*(asyncio.create_task(one(l, ip, p)) for (l, ip, p) in opens))
    finally:
        prog.finish()
        printer.cancel()
        try:
            await printer
        except asyncio.CancelledError:
            pass
    return results


# --------------------------------------------------------------------------- #
# Output
# --------------------------------------------------------------------------- #
def group_by_host(results):
    by = defaultdict(list)
    for r in results:
        by[(r["label"], r["ip"])].append(r)
    for recs in by.values():
        recs.sort(key=lambda r: r["port"])
    return dict(sorted(by.items(), key=lambda kv: (kv[0][0], kv[0][1])))


def write_outputs(results, opens, base, want_excel=True):
    written = []

    urls_path = base + ".urls.txt"
    with open(urls_path, "w", encoding="utf-8") as f:
        for r in sorted(results, key=lambda r: (r["label"], r["port"])):
            f.write(f"{r['scheme']}://{r['label']}:{r['port']}\n")
    written.append(urls_path)

    byhost_path = base + ".byhost.txt"
    grouped = group_by_host(results)
    with open(byhost_path, "w", encoding="utf-8") as f:
        for (label, ip), recs in grouped.items():
            ports = " ".join(f"{r['port']}/{r['scheme']}" for r in recs)
            shown = label if label == ip else f"{label} ({ip})"
            f.write(f"{shown}\t{ports}\n")
    written.append(byhost_path)

    open_path = base + ".open.txt"
    with open(open_path, "w", encoding="utf-8") as f:
        for label, ip, port in sorted(opens, key=lambda o: (o[1], o[2])):
            f.write(f"{ip}:{port}\n")
    written.append(open_path)

    jsonl_path = base + ".http.jsonl"
    with open(jsonl_path, "w", encoding="utf-8") as f:
        for r in results:
            f.write(json.dumps(r) + "\n")
    written.append(jsonl_path)

    csv_path = base + ".http.csv"
    cols = ["label", "ip", "port", "scheme", "url", "status", "server",
            "content_type", "title", "location"]
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for r in sorted(results, key=lambda r: (r["label"], r["port"])):
            w.writerow({k: r.get(k, "") for k in cols})
    written.append(csv_path)

    if want_excel:
        xlsx_path = base + ".xlsx"
        try:
            build_excel(results, xlsx_path)
            written.append(xlsx_path)
        except ImportError:
            sys.stderr.write("[warn] openpyxl not installed - skipping Excel export "
                             "(pip install openpyxl)\n")
    return written


def build_excel(results, path):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    https_fill = PatternFill("solid", fgColor="E2EFDA")

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "HTTP by host"
    ws.append(["Host", "IP", "# HTTP ports", "Ports", "URLs"])
    for c in ws[1]:
        c.fill, c.font = hdr_fill, hdr_font
    for (label, ip), recs in group_by_host(results).items():
        ports = ", ".join(f"{r['port']}/{r['scheme']}" for r in recs)
        urls = "\n".join(r["url"] for r in recs)
        ws.append([label, ip, len(recs), ports, urls])
    for i, w in enumerate((30, 16, 12, 42, 52), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        row[3].alignment = Alignment(wrap_text=True, vertical="top")
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Endpoints")
    cols = ["Host", "IP", "Port", "Scheme", "URL", "Status", "Server",
            "Content-Type", "Title", "Location"]
    ws2.append(cols)
    for c in ws2[1]:
        c.fill, c.font = hdr_fill, hdr_font
    for r in sorted(results, key=lambda r: (r["label"], r["port"])):
        ws2.append([r["label"], r["ip"], r["port"], r["scheme"], r["url"],
                    r["status"], r["server"], r["content_type"], r["title"], r["location"]])
        if r["scheme"] == "https":
            for c in ws2[ws2.max_row]:
                c.fill = https_fill
    for i, w in enumerate((28, 16, 7, 8, 46, 7, 24, 26, 36, 30), 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    wb.save(path)


def print_summary(results):
    grouped = group_by_host(results)
    print()
    print("=" * 64)
    print(f"HTTP ports found: {len(results)} endpoint(s) across {len(grouped)} host(s)")
    print("=" * 64)
    for (label, ip), recs in grouped.items():
        shown = label if label == ip else f"{label} ({ip})"
        ports = "  ".join(f"{r['port']}/{r['scheme']}" for r in recs)
        print(f"{shown:<32} {ports}")
    if not grouped:
        print("(none)")
    print()


# --------------------------------------------------------------------------- #
# Concurrency planning
# --------------------------------------------------------------------------- #
def dynamic_port_capacity():
    """Best-effort read of the Windows ephemeral port range; None elsewhere/on failure."""
    if not WIN:
        return None
    try:
        import subprocess
        out = subprocess.run(
            ["netsh", "int", "ipv4", "show", "dynamicport", "tcp"],
            capture_output=True, text=True, timeout=8,
        ).stdout
        m = re.search(r"Number of Ports\s*:\s*(\d+)", out)
        if m:
            return int(m.group(1))
    except Exception:
        pass
    return None


def plan_scan(args, total_pairs):
    """Decide (processes, per_process_concurrency, est_seconds, notes)."""
    cpu = os.cpu_count() or 1
    cap = dynamic_port_capacity()
    ephemeral = cap if cap else 60000
    max_procs = min(cpu, 16)
    notes = []

    if args.processes is not None:
        nproc = max(1, args.processes)
    elif args.budget_hours:
        need_cps = total_pairs / (args.budget_hours * 3600.0)
        nproc = max(1, min(max_procs, math.ceil(need_cps / PER_PROC_CPS)))
    else:
        nproc = min(8, cpu)

    if total_pairs < 20000:                       # spawn overhead not worth it
        nproc = 1

    conc = args.concurrency
    max_total = max(1000, ephemeral - 3000)       # keep total sockets within ephemeral range
    if nproc * conc > max_total:
        new = max(200, max_total // nproc)
        notes.append(f"per-process concurrency {conc:,} -> {new:,} to keep total sockets "
                     f"({nproc}x) within the ephemeral-port range")
        conc = new

    agg_cps = min(nproc * conc / args.connect_timeout, nproc * PER_PROC_CPS)
    est = total_pairs / agg_cps if agg_cps > 0 else 0.0
    if args.budget_hours and est > args.budget_hours * 3600 * 1.15:
        notes.append(f"note: this box is CPU-bound at ~{nproc * PER_PROC_CPS:,}/s aggregate; "
                     f"estimate ~{fmt_dur(est)} may exceed the {args.budget_hours}h budget "
                     f"(raise --processes if you have spare cores)")
    return nproc, conc, est, notes


def print_rate_tune(args):
    cap = dynamic_port_capacity()
    cpu = os.cpu_count() or 1
    print("Windows fast-scan tuning")
    print("------------------------")
    print(f"  CPU cores                     : {cpu}")
    if cap:
        print(f"  ephemeral TCP ports available : {cap:,}")
        print(f"  -> safe total sockets approx  : {max(1000, cap - 3000):,}")
    else:
        print("  ephemeral port range: (could not read via netsh)")
    print()
    print("  Speed model (important): one asyncio process is CPU-bound at")
    print(f"  ~{PER_PROC_CPS:,} connect-attempts/sec. Go faster by adding PROCESSES,")
    print("  not just concurrency. Aggregate rate ~= processes x that.")
    print(f"    e.g. 1200 * 65535 = 78.6M pairs / (8 procs * {PER_PROC_CPS:,}/s) ~= "
          f"{fmt_dur(1200*65535/(8*PER_PROC_CPS))}")
    print(f"    on your {cpu} cores you can push more:")
    print(f"      12 procs ~= {fmt_dur(1200*65535/(12*PER_PROC_CPS))},  "
          f"16 procs ~= {fmt_dur(1200*65535/(16*PER_PROC_CPS))}")
    print()
    print("  Use --budget-hours N to auto-pick the process count for a time target,")
    print("  or set -P/--processes directly (e.g. -P 12).")
    print()
    print("  Widen the ephemeral range (admin, only if total sockets are capped):")
    print("    netsh int ipv4 set dynamicport tcp start=1024 num=64511")
    print()
    print("  Shorten TIME_WAIT (admin, optional - this tool already RST-closes):")
    print("    reg add HKLM\\SYSTEM\\CurrentControlSet\\Services\\Tcpip\\Parameters "
          "/v TcpTimedWaitDelay /t REG_DWORD /d 30 /f")


# --------------------------------------------------------------------------- #
# Self-test
# --------------------------------------------------------------------------- #
def run_selftest():
    import http.server
    import tempfile
    import threading

    print("[selftest] starting local HTTP, HTTPS and silent-TCP servers...")

    class Quiet(http.server.BaseHTTPRequestHandler):
        def do_GET(self):
            body = b"<html><head><title>selftest ok</title></head><body>hi</body></html>"
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def log_message(self, *a):
            pass

    class QuietServer(http.server.HTTPServer):
        # phase 1 RST-closes probes; that surfaces as ConnectionResetError here - ignore it.
        def handle_error(self, request, client_address):
            if not isinstance(sys.exc_info()[1], (ConnectionResetError, ConnectionAbortedError)):
                super().handle_error(request, client_address)

    http_srv = QuietServer(("127.0.0.1", 0), Quiet)
    http_port = http_srv.server_address[1]
    threading.Thread(target=http_srv.serve_forever, daemon=True).start()

    # HTTPS server with a throwaway self-signed cert
    from cryptography import x509
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import rsa
    from cryptography.x509.oid import NameOID
    import datetime as _dt

    key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
    name = x509.Name([x509.NameAttribute(NameOID.COMMON_NAME, "localhost")])
    cert = (
        x509.CertificateBuilder()
        .subject_name(name).issuer_name(name)
        .public_key(key.public_key())
        .serial_number(x509.random_serial_number())
        .not_valid_before(_dt.datetime.now(_dt.timezone.utc) - _dt.timedelta(days=1))
        .not_valid_after(_dt.datetime.now(_dt.timezone.utc) + _dt.timedelta(days=1))
        .sign(key, hashes.SHA256())
    )
    tmp = tempfile.mkdtemp()
    cert_p, key_p = os.path.join(tmp, "c.pem"), os.path.join(tmp, "k.pem")
    with open(cert_p, "wb") as f:
        f.write(cert.public_bytes(serialization.Encoding.PEM))
    with open(key_p, "wb") as f:
        f.write(key.private_bytes(serialization.Encoding.PEM,
                                  serialization.PrivateFormat.TraditionalOpenSSL,
                                  serialization.NoEncryption()))
    https_srv = QuietServer(("127.0.0.1", 0), Quiet)
    https_port = https_srv.server_address[1]
    sctx = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
    sctx.load_cert_chain(cert_p, key_p)
    https_srv.socket = sctx.wrap_socket(https_srv.socket, server_side=True)
    threading.Thread(target=https_srv.serve_forever, daemon=True).start()

    # silent open TCP port (accepts, never speaks HTTP)
    silent = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    silent.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
    silent.bind(("127.0.0.1", 0))
    silent.listen(64)
    silent_port = silent.getsockname()[1]

    def _accept_loop():
        while True:
            try:
                c, _ = silent.accept()
                threading.Thread(target=lambda s: (time.sleep(2), s.close()),
                                 args=(c,), daemon=True).start()
            except OSError:
                return

    threading.Thread(target=_accept_loop, daemon=True).start()

    # a very likely-closed port
    probe = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    probe.bind(("127.0.0.1", 0))
    closed_port = probe.getsockname()[1]
    probe.close()

    ports = sorted({http_port, https_port, silent_port, closed_port})
    print(f"[selftest] http={http_port} https={https_port} silent={silent_port} "
          f"closed={closed_port}")

    cfg = {
        "concurrency": 100, "connect_timeout": 1.0, "linger": True,
        "pairs_per_batch": 1000, "quiet": True,
        "http_concurrency": 50, "http_connect_timeout": 3.0,
        "http_read_timeout": 3.0, "ua": DEFAULT_UA,
    }
    targets = [("127.0.0.1", "127.0.0.1")]

    async def _drive():
        journal = Journal("", enabled=False)
        opens = await scan_open_ports(targets, ports, cfg, journal, lambda opened=False: None)
        results = await identify_http(opens, cfg)
        return opens, results

    if WIN:
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    opens, results = asyncio.run(_drive())

    open_ports = {p for (_, _, p) in opens}
    http_ports = {r["port"]: r["scheme"] for r in results}

    ok = True

    def _check(cond, msg):
        nonlocal ok
        print(f"  [{'PASS' if cond else 'FAIL'}] {msg}")
        ok = ok and cond

    _check(http_port in open_ports, f"phase1 found HTTP port {http_port} open")
    _check(https_port in open_ports, f"phase1 found HTTPS port {https_port} open")
    _check(silent_port in open_ports, f"phase1 found silent port {silent_port} open")
    _check(closed_port not in open_ports, f"phase1 did NOT flag closed port {closed_port}")
    _check(http_ports.get(http_port) == "http", f"phase2 identified {http_port} as http")
    _check(http_ports.get(https_port) == "https", f"phase2 identified {https_port} as https")
    _check(silent_port not in http_ports, f"phase2 excluded non-HTTP silent port {silent_port}")

    http_srv.shutdown()
    https_srv.shutdown()
    silent.close()

    print()
    print("[selftest] " + ("ALL CHECKS PASSED" if ok else "FAILURES DETECTED"))
    return 0 if ok else 1


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def build_parser():
    p = argparse.ArgumentParser(
        description="masscan-style pure-Python port sweep -> HTTP identification.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("-t", "--targets", help="file with hosts (IP/CIDR/range/hostname, one per line)")
    p.add_argument("-p", "--ports", default="1-65535",
                   help="ports to scan: 'all' (default 1-65535), '80,443', '8000-9000'")
    p.add_argument("-c", "--concurrency", type=int, default=2000,
                   help="concurrent TCP connects PER PROCESS (default 2000); "
                        "total in-flight = processes x concurrency")
    p.add_argument("-P", "--processes", type=int, default=None,
                   help="worker processes for the sweep (default auto: min(8, cores); "
                        "1 for small jobs). One asyncio loop is CPU-bound, so this is "
                        "the main speed lever on multi-core boxes")
    p.add_argument("--budget-hours", type=float, default=None,
                   help="auto-pick the number of processes to finish the sweep in ~this many hours")
    p.add_argument("--connect-timeout", type=float, default=1.5,
                   help="TCP connect timeout in seconds for the sweep (default 1.5)")
    p.add_argument("--pairs-per-batch", type=int, default=20000,
                   help="host*port pairs scanned per journaled batch (default 20000)")
    p.add_argument("--no-linger", action="store_true",
                   help="do not RST-close open sockets (keep default TIME_WAIT behavior)")
    p.add_argument("--open-only", action="store_true",
                   help="only run the port sweep; skip HTTP identification")
    p.add_argument("--http-concurrency", type=int, default=200,
                   help="concurrency for the HTTP identification phase (default 200)")
    p.add_argument("--http-connect-timeout", type=float, default=4.0)
    p.add_argument("--http-read-timeout", type=float, default=4.0)
    p.add_argument("--ua", default=DEFAULT_UA, help="User-Agent for HTTP probes")
    p.add_argument("-o", "--output", default=None,
                   help="output base path (default scan_<timestamp>)")
    p.add_argument("--no-excel", action="store_true", help="skip the .xlsx export")
    p.add_argument("--resume", action="store_true",
                   help="resume from <output>.journal.jsonl (skip completed ports)")
    p.add_argument("-q", "--quiet", action="store_true", help="suppress the live progress line")
    p.add_argument("--rate-tune", action="store_true",
                   help="print Windows tuning advice and exit")
    p.add_argument("--selftest", action="store_true",
                   help="run a local end-to-end self-test and exit")
    return p


def main(argv=None):
    parser = build_parser()
    argv = list(sys.argv[1:] if argv is None else argv)
    args = parser.parse_args(argv)

    if args.rate_tune:
        print_rate_tune(args)
        return 0
    if args.selftest:
        return run_selftest()
    if not args.targets:
        parser.error("-t/--targets is required (or use --selftest / --rate-tune)")

    with open(args.targets, encoding="utf-8") as f:
        raw = expand_targets(f.readlines())
    if not raw:
        sys.stderr.write("[error] no valid targets\n")
        return 2
    print(f"[+] expanding targets... ", end="", flush=True)
    targets = resolve_targets(raw)
    ports = parse_ports(args.ports)
    print(f"{len(targets)} host(s), {len(ports)} port(s) -> {len(targets)*len(ports):,} pairs")
    if not targets:
        sys.stderr.write("[error] no resolvable targets\n")
        return 2

    base = args.output or datetime.now().strftime("scan_%Y%m%d_%H%M%S")

    done_ports, resumed_opens = (set(), [])
    if args.resume:
        done_ports, resumed_opens = load_all_journals(base)
        if done_ports or resumed_opens:
            print(f"[+] resume: {len(done_ports)} port(s) already done, "
                  f"{len(resumed_opens)} open port(s) recovered")

    scan_ports = [p for p in ports if p not in done_ports]
    total_pairs = len(targets) * len(scan_ports)

    nproc, concurrency, est, notes = plan_scan(args, total_pairs)
    for n in notes:
        print(f"[i] {n}")

    print(f"[+] sweep: processes={nproc}  concurrency/proc={concurrency:,}  "
          f"total_sockets<={nproc*concurrency:,}  connect_timeout={args.connect_timeout}s  "
          f"linger={'off' if args.no_linger else 'on(RST)'}")
    print(f"[+] estimated sweep time: ~{fmt_dur(est)}  "
          f"(faster if hosts RST closed ports; slower if heavily firewalled)")

    cfg = {
        "concurrency": concurrency,
        "connect_timeout": args.connect_timeout,
        "linger": not args.no_linger,
        "pairs_per_batch": args.pairs_per_batch,
        "quiet": args.quiet,
        "journal": True,
        "http_concurrency": args.http_concurrency,
        "http_connect_timeout": args.http_connect_timeout,
        "http_read_timeout": args.http_read_timeout,
        "ua": args.ua,
    }

    if WIN:
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

    opens = list(resumed_opens)
    t0 = time.time()
    try:
        if not scan_ports:
            print("[i] all ports already scanned (resume) - skipping sweep")
        elif nproc <= 1:
            journal = Journal(base + ".journal.jsonl", enabled=True)
            try:
                opens.extend(run_single(targets, scan_ports, cfg, journal))
            finally:
                journal.close()
        else:
            opens.extend(run_multi(targets, scan_ports, cfg, base, nproc))
    except KeyboardInterrupt:
        print("\n[!] interrupted - partial results kept (use --resume to continue)")

    # dedup opens
    opens = sorted(set(opens), key=lambda o: (o[1], o[2]))
    print(f"[+] sweep complete: {len(opens)} open port(s) in {fmt_dur(time.time()-t0)}")

    results = []
    if args.open_only:
        print("[i] --open-only: skipping HTTP identification")
    elif opens:
        results = asyncio.run(identify_http(opens, cfg))

    written = write_outputs(results, opens, base, want_excel=not args.no_excel)
    print_summary(results)
    print("[+] wrote:")
    for w in written:
        print(f"    {w}")
    print(f"\n[+] feed the URLs straight into the auth recon:")
    print(f"    python api_auth_recon.py --targets {base}.urls.txt --current-user")
    return 0


if __name__ == "__main__":
    sys.exit(main())
