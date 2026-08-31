#!/usr/bin/env python3
"""
api_auth_recon.py  -  Authorized API authentication recon for an internal network.

Pipeline, per host:
  1. Probe common HTTP(S) ports  -> find live web services
  2. Look for Swagger / OpenAPI specs on common paths
  3. Parse each spec, extract GET operations
  4. Send UNAUTHENTICATED GET requests and classify the auth behaviour

Output: a live console view + an Excel workbook (3 sheets: Web, Swagger, Endpoints).

Crash-safety:
  - every finding is appended to a JSONL journal and flushed immediately, so even
    a hard kill (SIGKILL) leaves everything on disk;
  - the Excel file is rebuilt from the journal periodically, on Ctrl-C, and at exit,
    written atomically (temp file -> os.replace);
  - re-running with the same --out RESUMES: hosts already finished are skipped.

Only idempotent GET requests are sent. Paths that look state-changing are skipped.

NOTE (Windows domain-joined): the `requests` library does NOT auto-send your
Kerberos/NTLM credentials the way a browser / WinHTTP would, so by default these
results reflect TRUE anonymous access -- which is exactly what this test wants.

--current-user adds a SECOND, opt-in check on top of that: for each endpoint that
came back PROTECTED behind NTLM/Negotiate, it retries once as the current
logged-in Windows user (SSPI, no password) and records whether it opens up, in a
separate 'auth_as_user' column. The anonymous verdict is never overwritten, so
you keep both signals: what is open to anyone, and what any domain user can reach.
Needs the requests-negotiate-sspi package (Windows only).
"""

import argparse
import atexit
import json
import os
import re
import signal
import socket
import sys
import threading
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from urllib.parse import urljoin, urlparse

import requests
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from rich.console import Console
from rich.progress import (BarColumn, MofNCompleteColumn, Progress, TextColumn,
                           TimeElapsedColumn)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
console = Console()

# ----------------------------------------------------------------------------- config / defaults
DEFAULT_PORTS = [80, 443, 8080, 8443, 8000, 8888, 5000, 3000,
                 9000, 9090, 8081, 7001, 9443, 10443]
HTTPS_PORTS = {443, 8443, 9443, 10443, 4443}

SWAGGER_PATHS = [
    "/swagger/v1/swagger.json", "/swagger/v2/swagger.json",
    "/v2/api-docs", "/v3/api-docs", "/openapi.json", "/openapi.yaml",
    "/api-docs", "/api/swagger.json", "/swagger.json",
    "/swagger-ui.html", "/swagger-ui/index.html", "/swagger/index.html",
    "/api/v1/openapi.json", "/api/v2/openapi.json", "/docs/swagger.json",
    "/api-docs/swagger.json", "/swagger-resources",
]

# path fragments that suggest a GET could still have side effects -> never touch
STATECHANGE_MARKERS = ("delete", "remove", "destroy", "drop", "reset", "revoke",
                       "shutdown", "reboot", "restart", "kill", "purge", "wipe",
                       "logout", "signout", "deactivate", "disable", "terminate",
                       "export", "download", "backup", "restore", "flush")

# body markers that mean "200 but actually not authenticated"
AUTH_BODY_MARKERS = ("unauthorized", "not authenticated", "authentication required",
                     "access denied", "please log in", "please sign in",
                     "login required", "invalid token", "missing token",
                     "forbidden", "session expired")

TITLE_RE = re.compile(r"<title[^>]*>(.*?)</title>", re.I | re.S)


# ----------------------------------------------------------------------------- journal / excel writer
class Journal:
    """Append-only crash-safe store. Records are dicts with a 'type' field."""

    def __init__(self, path):
        self.path = path
        self.xlsx = os.path.splitext(path)[0] + ".xlsx" if path.endswith(".jsonl") \
            else path + ".jsonl_notused"
        self._lock = threading.Lock()
        self._fh = open(path, "a", encoding="utf-8", buffering=1)  # line-buffered
        self._records = []          # in-memory mirror for excel building
        self._load_existing()

    def _load_existing(self):
        # read anything already there (resume) into memory
        try:
            with open(self.path, "r", encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        self._records.append(json.loads(line))
                    except json.JSONDecodeError:
                        pass
        except FileNotFoundError:
            pass

    def done_hosts(self):
        return {r["host"] for r in self._records if r.get("type") == "host_done"}

    def write(self, rec):
        rec.setdefault("ts", datetime.now().isoformat(timespec="seconds"))
        line = json.dumps(rec, ensure_ascii=False)
        with self._lock:
            self._fh.write(line + "\n")
            self._fh.flush()
            os.fsync(self._fh.fileno())
            self._records.append(rec)

    def snapshot(self):
        with self._lock:
            recs = list(self._records)
        return recs

    def close(self):
        try:
            self._fh.close()
        except Exception:
            pass


def build_excel(records, out_xlsx):
    """Rebuild the whole workbook from journal records. Atomic write."""
    wb = Workbook()

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    unauth_fill = PatternFill("solid", fgColor="FFC7CE")   # red-ish
    unauth_font = Font(color="9C0006", bold=True)
    warn_fill = PatternFill("solid", fgColor="FFEB9C")     # amber
    open_fill = PatternFill("solid", fgColor="C6EFCE")     # green (locked / denied-as-user)
    open_font = Font(color="006100", bold=True)
    esc_fill = PatternFill("solid", fgColor="E4C6F0")      # purple (domain-user escalation)
    esc_font = Font(color="6B1F8A", bold=True)

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    def autosize(ws, maxw=80):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, maxw)

    # ---- Web services sheet
    ws = wb.active
    ws.title = "Web Services"
    web_cols = ["host", "port", "scheme", "url", "status", "server", "title", "ts"]
    ws.append(web_cols)
    for r in records:
        if r.get("type") == "web":
            ws.append([r.get(k, "") for k in web_cols])
    style_header(ws, len(web_cols))
    autosize(ws)

    # ---- Swagger sheet
    ws = wb.create_sheet("Swagger")
    sw_cols = ["service_url", "swagger_url", "spec_title", "spec_version",
               "get_endpoints", "ts"]
    ws.append(sw_cols)
    for r in records:
        if r.get("type") == "swagger":
            ws.append([r.get(k, "") for k in sw_cols])
    style_header(ws, len(sw_cols))
    autosize(ws)

    # ---- Endpoints sheet (the money sheet)
    ws = wb.create_sheet("Endpoints")
    ep_cols = ["verdict", "auth_as_user", "access", "method", "url", "status",
               "www_authenticate", "location", "content_type", "resp_len",
               "swagger_url", "ts"]
    ws.append(ep_cols)
    row_i = 1
    for r in records:
        if r.get("type") != "endpoint":
            continue
        row_i += 1
        ws.append([r.get(k, "") for k in ep_cols])
        verdict = r.get("verdict", "")
        access = r.get("access", "")
        if verdict == "UNAUTHENTICATED":
            for c in range(1, len(ep_cols) + 1):
                ws.cell(row=row_i, column=c).fill = unauth_fill
            ws.cell(row=row_i, column=1).font = unauth_font
        elif verdict in ("200-AUTH-BODY", "OTHER", "ERROR"):
            ws.cell(row=row_i, column=1).fill = warn_fill
        # access-matrix column (col 3): flag the escalation, mark the locked ones
        if access == "DOMAIN-USER":
            for c in range(1, len(ep_cols) + 1):
                ws.cell(row=row_i, column=c).fill = esc_fill
            ws.cell(row=row_i, column=3).font = esc_font
        elif access == "LOCKED":
            ws.cell(row=row_i, column=3).fill = open_fill
            ws.cell(row=row_i, column=3).font = open_font
    style_header(ws, len(ep_cols))
    autosize(ws)

    # auto-filter on endpoints
    ws.auto_filter.ref = ws.dimensions

    # ---- Access Matrix sheet (only when a --current-user pass produced authed results)
    authed_recs = [r for r in records
                   if r.get("type") == "endpoint" and r.get("auth_as_user")]
    if authed_recs:
        ws = wb.create_sheet("Access Matrix")
        anon_axis = sorted({r.get("verdict", "") for r in authed_recs})
        authed_axis = sorted({r.get("auth_as_user", "") for r in authed_recs})
        pair = Counter((r.get("verdict", ""), r.get("auth_as_user", ""))
                       for r in authed_recs)

        # cross-tab: anonymous verdict (rows) x current-user verdict (cols)
        ws.append(["anonymous \\ current-user"] + authed_axis + ["row total"])
        for a in anon_axis:
            row = [a] + [pair.get((a, b), 0) for b in authed_axis]
            row.append(sum(pair.get((a, b), 0) for b in authed_axis))
            ws.append(row)
        ws.append(["col total"]
                  + [sum(pair.get((a, b), 0) for a in anon_axis) for b in authed_axis]
                  + [len(authed_recs)])
        style_header(ws, len(authed_axis) + 2)
        for r_i in range(2, 3 + len(anon_axis)):
            ws.cell(row=r_i, column=1).font = Font(bold=True)

        # access-class rollup underneath the cross-tab
        classes = Counter(r.get("access", "") for r in authed_recs)
        meaning = {
            "OPEN-TO-ALL": "reachable with NO credentials -- top priority",
            "DOMAIN-USER": "denied anonymously, opens to any logged-in domain user",
            "LOCKED": "denied both anonymously and as the current user",
            "OTHER": "redirects / errors / mixed -- check manually",
        }
        cls_fill = {"OPEN-TO-ALL": unauth_fill, "DOMAIN-USER": esc_fill,
                    "LOCKED": open_fill}
        start = len(anon_axis) + 4
        for j, label in enumerate(("Access class", "count", "meaning"), start=1):
            ws.cell(row=start, column=j, value=label).font = Font(bold=True)
        ri = start
        for cls in ("OPEN-TO-ALL", "DOMAIN-USER", "LOCKED", "OTHER"):
            if not classes.get(cls):
                continue
            ri += 1
            ws.cell(row=ri, column=1, value=cls)
            ws.cell(row=ri, column=2, value=classes[cls])
            ws.cell(row=ri, column=3, value=meaning[cls])
            if cls in cls_fill:
                for c in (1, 2, 3):
                    ws.cell(row=ri, column=c).fill = cls_fill[cls]
        autosize(ws)

    tmp = out_xlsx + ".tmp"
    wb.save(tmp)
    os.replace(tmp, out_xlsx)


# ----------------------------------------------------------------------------- http helpers
def make_session(ua, retries=0):
    s = requests.Session()
    s.headers["User-Agent"] = ua
    adapter = HTTPAdapter(pool_connections=50, pool_maxsize=50, max_retries=retries)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    return s


def tcp_open(host, port, timeout=1.5):
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except Exception:
        return False


def probe_http(session, host, port, timeout):
    """Return dict describing a live web service on host:port, or None."""
    if not tcp_open(host, port, timeout=min(timeout, 2.0)):
        return None
    schemes = (["https", "http"] if port in HTTPS_PORTS else ["http", "https"])
    for scheme in schemes:
        url = f"{scheme}://{host}:{port}/"
        try:
            r = session.get(url, timeout=timeout, verify=False, allow_redirects=True)
        except requests.exceptions.SSLError:
            continue          # wrong scheme, try the other
        except requests.RequestException:
            continue
        title = ""
        m = TITLE_RE.search(r.text[:4096]) if r.text else None
        if m:
            title = re.sub(r"\s+", " ", m.group(1)).strip()[:120]
        return {
            "host": host, "port": port, "scheme": scheme,
            "url": f"{scheme}://{host}:{port}",
            "status": r.status_code,
            "server": r.headers.get("Server", "")[:80],
            "title": title,
        }
    return None


def find_swaggers(session, base_url, timeout):
    """Try common swagger paths under base_url. Return list of (url, spec_dict)."""
    found = []
    seen_specs = set()
    for p in SWAGGER_PATHS:
        url = base_url.rstrip("/") + p
        try:
            r = session.get(url, timeout=timeout, verify=False, allow_redirects=True)
        except requests.RequestException:
            continue
        if r.status_code != 200:
            continue
        ctype = r.headers.get("Content-Type", "").lower()
        spec = None
        # JSON specs
        if "json" in ctype or r.text.lstrip().startswith("{"):
            try:
                spec = r.json()
            except ValueError:
                spec = None
        if isinstance(spec, dict) and ("swagger" in spec or "openapi" in spec or "paths" in spec):
            key = json.dumps(sorted(spec.get("paths", {}).keys()))[:2000]
            if key not in seen_specs:
                seen_specs.add(key)
                found.append((r.url, spec))
    return found


def spec_base_urls(spec, spec_url):
    """Resolve the base URL(s) the endpoints hang off, per OpenAPI3 / Swagger2."""
    parsed = urlparse(spec_url)
    origin = f"{parsed.scheme}://{parsed.netloc}"
    bases = []
    if "servers" in spec and isinstance(spec["servers"], list) and spec["servers"]:
        for srv in spec["servers"]:
            u = srv.get("url", "/") if isinstance(srv, dict) else str(srv)
            # substitute variable defaults {var}
            for var, meta in (srv.get("variables", {}) or {}).items() if isinstance(srv, dict) else []:
                u = u.replace("{" + var + "}", str(meta.get("default", "")))
            if not u.startswith("http"):
                u = urljoin(spec_url, u)
            bases.append(u.rstrip("/"))
    elif "swagger" in spec:   # v2
        schemes = spec.get("schemes") or [parsed.scheme]
        host = spec.get("host") or parsed.netloc
        base_path = spec.get("basePath", "") or ""
        for sc in schemes:
            bases.append(f"{sc}://{host}{base_path}".rstrip("/"))
    if not bases:
        bases = [origin]
    # dedup, keep order
    out, seen = [], set()
    for b in bases:
        if b not in seen:
            seen.add(b)
            out.append(b)
    return out


def extract_get_endpoints(spec):
    """Return list of raw path strings that expose a GET operation."""
    eps = []
    for path, ops in (spec.get("paths") or {}).items():
        if not isinstance(ops, dict):
            continue
        if any(k.lower() == "get" for k in ops.keys()):
            eps.append(path)
    return eps


def substitute_params(path):
    # replace {id}, {userId}, : style params with a benign value
    path = re.sub(r"\{[^}]+\}", "1", path)
    path = re.sub(r":([A-Za-z_][A-Za-z0-9_]*)", "1", path)
    return path


def looks_state_changing(path):
    low = path.lower()
    return any(m in low for m in STATECHANGE_MARKERS)


def classify(resp):
    """Return (verdict, detail_dict)."""
    sc = resp.status_code
    www = resp.headers.get("WWW-Authenticate", "")
    loc = resp.headers.get("Location", "")
    ctype = resp.headers.get("Content-Type", "")[:60]
    body = (resp.text or "")[:2048].lower()
    detail = {"status": sc, "www_authenticate": www[:80], "location": loc[:160],
              "content_type": ctype, "resp_len": len(resp.content or b"")}

    if sc in (401, 403):
        return "PROTECTED", detail
    if www:
        return "PROTECTED", detail
    if sc in (301, 302, 303, 307, 308):
        if re.search(r"(login|signin|sso|adfs|oauth|auth|saml|openid)", loc, re.I):
            return "REDIRECT-SSO", detail
        return "REDIRECT", detail
    if sc == 200:
        if any(m in body for m in AUTH_BODY_MARKERS):
            return "200-AUTH-BODY", detail
        return "UNAUTHENTICATED", detail
    return "OTHER", detail


# ----------------------------------------------------------------------------- current-user auth
# Normalised authed verdicts (short, since they head the access-matrix columns).
_AUTHED_NAMES = {"UNAUTHENTICATED": "OPEN", "200-AUTH-BODY": "AUTH-BODY",
                 "PROTECTED": "DENIED"}


def _negotiate_auth():
    """A fresh SSPI Negotiate/NTLM auth handler bound to the current logon session.

    New instance per request: the handler carries per-exchange state, so sharing
    one across the worker threads would be unsafe.
    """
    from requests_negotiate_sspi import HttpNegotiateAuth
    return HttpNegotiateAuth()


def probe_as_current_user(session, url, timeout):
    """GET url as the current Windows user (SSPI). Return a short authed verdict."""
    try:
        auth = _negotiate_auth()
    except Exception:
        return "no-sspi"
    try:
        r = session.get(url, timeout=timeout, verify=False,
                        allow_redirects=False, auth=auth)
    except requests.RequestException:
        return "error"
    v, _ = classify(r)
    return _AUTHED_NAMES.get(v, v)


def access_class(anon, authed):
    """Collapse an (anonymous, current-user) verdict pair into one access class."""
    if anon == "UNAUTHENTICATED":
        return "OPEN-TO-ALL"          # reachable with no credentials at all
    if authed == "OPEN":
        return "DOMAIN-USER"          # denied anon, opens to any logged-in user
    if (anon in ("PROTECTED", "REDIRECT-SSO", "200-AUTH-BODY")
            and authed in ("DENIED", "AUTH-BODY", "REDIRECT-SSO")):
        return "LOCKED"               # denied both ways
    return "OTHER"                    # redirects / errors / mixed


# ----------------------------------------------------------------------------- per-host worker
def process_host(target, journal, cfg):
    """Full chain for one host. Emits records as it goes, then host_done."""
    host = target["host"]
    ports = target["ports"]
    session = make_session(cfg["ua"], retries=cfg["retries"])
    timeout = cfg["timeout"]
    delay = cfg["delay"]

    web_services = []
    for port in ports:
        svc = probe_http(session, host, port, timeout)
        if svc:
            journal.write({"type": "web", **svc})
            console.print(
                f"[green][live][/green] {svc['url']} -> {svc['status']} "
                f"[dim]{svc['server']}[/dim] "
                + (f'"{svc["title"]}"' if svc["title"] else ""))
            web_services.append(svc)
        if delay:
            time.sleep(delay)

    for svc in web_services:
        base = svc["url"]
        swaggers = find_swaggers(session, base, timeout)
        for swurl, spec in swaggers:
            get_paths = extract_get_endpoints(spec)
            journal.write({
                "type": "swagger", "service_url": base, "swagger_url": swurl,
                "spec_title": (spec.get("info", {}) or {}).get("title", "")[:120],
                "spec_version": (spec.get("info", {}) or {}).get("version", ""),
                "get_endpoints": len(get_paths),
            })
            console.print(
                f"[cyan][swagger][/cyan] {swurl} "
                f"[dim]({len(get_paths)} GET endpoints)[/dim]")

            bases = spec_base_urls(spec, swurl)
            api_base = bases[0]
            # optionally sample
            paths = get_paths
            if cfg["per_api"] and cfg["per_api"] > 0:
                paths = paths[:cfg["per_api"]]

            for raw in paths:
                if looks_state_changing(raw):
                    continue
                url = api_base + substitute_params(raw)
                try:
                    r = session.get(url, timeout=timeout, verify=False,
                                    allow_redirects=False)
                except requests.RequestException as e:
                    journal.write({"type": "endpoint", "verdict": "ERROR",
                                   "auth_as_user": "", "access": "",
                                   "method": "GET", "url": url, "status": "",
                                   "www_authenticate": "", "location": "",
                                   "content_type": "", "resp_len": 0,
                                   "swagger_url": swurl, "error": str(e)[:120]})
                    if delay:
                        time.sleep(delay)
                    continue
                verdict, detail = classify(r)

                # opt-in second pass: same endpoint as the current Windows user,
                # so every row carries both an anonymous and an authed result.
                authed, access = "", ""
                if cfg["current_user"]:
                    if delay:
                        time.sleep(delay)
                    authed = probe_as_current_user(session, url, timeout)
                    access = access_class(verdict, authed)

                journal.write({"type": "endpoint", "verdict": verdict,
                               "auth_as_user": authed, "access": access,
                               "method": "GET", "url": url,
                               "swagger_url": swurl, **detail})
                if verdict == "UNAUTHENTICATED":
                    console.print(
                        f"  [bold red][UNAUTH][/bold red] GET {url} -> "
                        f"{detail['status']}  [red]<-- no auth[/red]")
                elif access == "DOMAIN-USER":
                    scheme = detail["www_authenticate"].split(" ")[0] if detail["www_authenticate"] else ""
                    console.print(
                        f"  [bold magenta][DOMAIN-USER][/bold magenta] GET {url} -> "
                        f"anon {detail['status']} {scheme} "
                        f"[magenta]<-- opens to current user[/magenta]")
                elif verdict == "PROTECTED":
                    scheme = detail["www_authenticate"].split(" ")[0] if detail["www_authenticate"] else ""
                    tail = f" [dim](as-user: {authed})[/dim]" if authed else ""
                    console.print(
                        f"  [dim][ok] GET {url} -> {detail['status']} {scheme}[/dim]{tail}")
                elif verdict.startswith("REDIRECT"):
                    console.print(
                        f"  [yellow][redir][/yellow] GET {url} -> {detail['status']} "
                        f"[dim]{detail['location'][:60]}[/dim]")
                else:
                    console.print(
                        f"  [dim][{verdict}] GET {url} -> {detail['status']}[/dim]")
                if delay:
                    time.sleep(delay)

    journal.write({"type": "host_done", "host": host})
    session.close()
    return host


# ----------------------------------------------------------------------------- input parsing
def read_targets(path, default_ports):
    targets = {}
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            for tok in re.split(r"[,\s]+", line):
                if not tok:
                    continue
                if ":" in tok and not tok.count(":") > 1:  # host:port (skip IPv6-ish)
                    host, _, p = tok.partition(":")
                    try:
                        port = int(p)
                    except ValueError:
                        host, port = tok, None
                    if port:
                        targets.setdefault(host, set()).add(port)
                        continue
                targets.setdefault(tok, set())
    # fill defaults for hosts with no explicit port
    out = []
    for host, ports in targets.items():
        out.append({"host": host, "ports": sorted(ports) if ports else list(default_ports)})
    return out


# ----------------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Authorized internal API auth recon.")
    ap.add_argument("-i", "--input", help="file with hosts (host or host:port per line/comma)")
    ap.add_argument("-o", "--out", default="api_recon.jsonl",
                    help="journal path (.jsonl). Excel is <out>.xlsx (default: api_recon.jsonl)")
    ap.add_argument("-w", "--workers", type=int, default=20, help="concurrent hosts (default 20)")
    ap.add_argument("-t", "--timeout", type=float, default=7.0, help="per-request timeout s (default 7)")
    ap.add_argument("--ports", help="comma list of ports to override defaults")
    ap.add_argument("--per-api", type=int, default=0,
                    help="max GET endpoints tested per swagger (0 = all, default 0)")
    ap.add_argument("--delay", type=float, default=0.0, help="delay between requests s (rate limit)")
    ap.add_argument("--retries", type=int, default=0, help="request retries (default 0)")
    ap.add_argument("--current-user", action="store_true",
                    help="also test every endpoint as the current logged-in Windows "
                         "user (SSPI Negotiate/NTLM, no password) and build an access "
                         "matrix of anonymous vs current-user access. Doubles the "
                         "request count. Needs requests-negotiate-sspi (Windows)")
    ap.add_argument("--ua", default="API-Auth-Recon/1.0 (authorized assessment)",
                    help="User-Agent")
    ap.add_argument("--snapshot-every", type=int, default=20,
                    help="rebuild Excel every N finished hosts (default 20)")
    ap.add_argument("--rebuild", action="store_true",
                    help="just rebuild the .xlsx from an existing journal and exit")
    args = ap.parse_args()

    out_jsonl = args.out if args.out.endswith(".jsonl") else args.out + ".jsonl"
    out_xlsx = os.path.splitext(out_jsonl)[0] + ".xlsx"

    journal = Journal(out_jsonl)

    if args.rebuild:
        build_excel(journal.snapshot(), out_xlsx)
        console.print(f"[green]Rebuilt[/green] {out_xlsx} from {out_jsonl}")
        journal.close()
        return

    if not args.input:
        ap.error("--input is required (unless --rebuild)")

    if args.current_user:
        try:
            _negotiate_auth()   # fail fast with a clear message if the backend is missing
        except Exception as e:
            console.print("[red]--current-user needs the SSPI backend, which failed to "
                          f"load:[/red] {e}")
            console.print("[dim]Install it on Windows with:  pip install "
                          "requests-negotiate-sspi[/dim]")
            journal.close()
            return
        console.print("[cyan]current-user mode:[/cyan] each endpoint is checked "
                      "anonymously AND as the current Windows user; the report gets an "
                      "Access Matrix sheet. [dim](this doubles the request count)[/dim]")

    default_ports = ([int(p) for p in args.ports.split(",")] if args.ports else DEFAULT_PORTS)
    targets = read_targets(args.input, default_ports)

    done = journal.done_hosts()
    pending = [t for t in targets if t["host"] not in done]
    console.print(
        f"[bold]{len(targets)}[/bold] hosts total, "
        f"[green]{len(done)}[/green] already done, "
        f"[yellow]{len(pending)}[/yellow] to scan. "
        f"Journal: {out_jsonl}  Excel: {out_xlsx}")
    console.print("[dim]Ctrl-C saves everything to Excel and exits. "
                  "Re-run same command to resume.[/dim]\n")

    cfg = {"ua": args.ua, "timeout": args.timeout, "delay": args.delay,
           "retries": args.retries, "per_api": args.per_api,
           "current_user": args.current_user}

    # ------- finalize hooks (signal / atexit) so Excel is always saved
    _finalized = threading.Event()

    def finalize(*_):
        if _finalized.is_set():
            return
        _finalized.set()
        try:
            build_excel(journal.snapshot(), out_xlsx)
            console.print(f"\n[bold green]Saved Excel:[/bold green] {out_xlsx}")
        except Exception as e:
            console.print(f"[red]Excel build failed:[/red] {e} "
                          f"(journal is intact at {out_jsonl}, run --rebuild)")
        finally:
            journal.close()

    atexit.register(finalize)
    try:
        signal.signal(signal.SIGTERM, lambda *a: (finalize(), sys.exit(0)))
    except Exception:
        pass

    if not pending:
        console.print("[green]Nothing to do.[/green]")
        finalize()
        return

    progress = Progress(
        TextColumn("[progress.description]{task.description}"),
        BarColumn(), MofNCompleteColumn(), TimeElapsedColumn(),
        console=console, transient=False)
    completed = 0
    stop = threading.Event()

    with progress:
        task = progress.add_task("[cyan]hosts", total=len(pending))
        with ThreadPoolExecutor(max_workers=args.workers) as ex:
            futures = {ex.submit(process_host, t, journal, cfg): t for t in pending}
            try:
                for fut in as_completed(futures):
                    if stop.is_set():
                        break
                    host = futures[fut]["host"]
                    try:
                        fut.result()
                    except Exception as e:
                        console.print(f"[red]host {host} errored:[/red] {e}")
                        journal.write({"type": "host_done", "host": host,
                                       "error": str(e)[:200]})
                    completed += 1
                    progress.advance(task)
                    if args.snapshot_every and completed % args.snapshot_every == 0:
                        build_excel(journal.snapshot(), out_xlsx)
            except KeyboardInterrupt:
                console.print("\n[yellow]Ctrl-C -- stopping, saving Excel...[/yellow]")
                stop.set()
                for f in futures:
                    f.cancel()
                # let running workers finish their current write; pool __exit__ waits

    # summary
    recs = journal.snapshot()
    n_web = sum(1 for r in recs if r.get("type") == "web")
    n_sw = sum(1 for r in recs if r.get("type") == "swagger")
    n_un = sum(1 for r in recs if r.get("type") == "endpoint" and r.get("verdict") == "UNAUTHENTICATED")
    n_ep = sum(1 for r in recs if r.get("type") == "endpoint")
    finalize()
    summary = (f"\n[bold]Summary:[/bold] {n_web} web services, {n_sw} swagger specs, "
               f"{n_ep} GET endpoints tested, "
               f"[bold red]{n_un} UNAUTHENTICATED[/bold red]")
    if args.current_user:
        n_du = sum(1 for r in recs if r.get("type") == "endpoint"
                   and r.get("access") == "DOMAIN-USER")
        summary += (f", [bold magenta]{n_du} open to the current domain user"
                    f"[/bold magenta] (see the Access Matrix sheet)")
    console.print(summary + ".")


if __name__ == "__main__":
    main()
